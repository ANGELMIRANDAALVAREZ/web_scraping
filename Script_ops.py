from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from tqdm import tqdm
import time

def get_image_counter_selenium(url, driver, attempt=1):
    try:
        # Ir a la URL
        driver.get(url)

        # Esperar a que la página inicial se cargue completamente
        wait = WebDriverWait(driver, 40)
        wait.until(lambda d: d.execute_script('return document.readyState') == 'complete')
        driver.execute_script("document.body.style.zoom='5%'")

        # En el segundo intento, desplazarse al final de la página
        if attempt == 2:
            total_height = driver.execute_script("return document.body.scrollHeight")
            driver.execute_script(f"window.scrollTo(0, {total_height});")
            driver.execute_script("document.body.style.zoom='5%'")
            time.sleep(15)

        # Esperar un tiempo adicional para que los elementos se carguen después del desplazamiento
        time.sleep(5)  # Espera 5 segundos

        image_counters = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'span.label')))

        # Asegurarse de que se encontraron suficientes elementos
        if len(image_counters) < 2:
            print("No se encontraron suficientes elementos 'span.label'.")
            return "sin dimples"
            raise Exception("Elementos insuficientes")  # Forzar reintento

        # Obtener el texto del segundo elemento
        return image_counters[1].get_attribute('textContent').strip()
    except Exception as e:
        print(f"Error al obtener contador de imágenes para la URL {url}: {e}")
        if attempt < 2:
            print("Reintentando...")
            return get_image_counter_selenium(url, driver, attempt + 1)
        return 'Error al obtener el contador'
def process_urls(start_row, end_row, file_path):
    chrome_options = Options()
    #chrome_options.add_argument("--headless") #  ejecutar sin abrir el navegador
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in tqdm(range(start_row, end_row + 1), desc=f"Procesando filas {start_row} a {end_row}"):
        url = sheet[f'A{row}'].value
        image_counter_text = get_image_counter_selenium(url, driver)

        if image_counter_text == 'Error al obtener el contador':
            sheet[f'D{row}'] = "sin dimples"
        else:
            sheet[f'D{row}'] = image_counter_text
        #guardar el progreso cada # determinado de URL´s:
        if (row - start_row+1)%1==0:
            workbook.save(file_path)
            print (f"progreso guardado en la fila {row}.")

    driver.quit()
    workbook.save(file_path)
    print(f"Proceso completado para filas {start_row} a {end_row}.")

def main():
    file_path = '/Users/angel1/Desktop/Ops.xlsx'
    total_rows = 126148  # Total de URLs
    threads = 2  # Número de hilos
    rows_per_thread = total_rows // threads

    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = []
        for i in range(threads):
            #start_row = i * rows_per_thread + 1
            start_row = 1
            end_row = (i + 1) * rows_per_thread
            #i = 900
            if i == threads - 1:  # Ajuste para el último segmento
                end_row = total_rows
            futures.append(executor.submit(process_urls, start_row, end_row, file_path))

        # Esperar a que todos los hilos terminen
        for future in futures:
            future.result()

if __name__ == "__main__":
    main()
chrome_options = Options()
#chrome_options.add_argument("--headless") #  ejecutar sin abrir el navegador

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
workbook = openpyxl.load_workbook('/Users/angel1/Desktop/Ops.xlsx')
sheet = workbook.active

for row in tqdm(range(1, sheet.max_row + 1), desc="Obteniendo contadores de imágenes"):
    url = sheet[f'A{row}'].value
    image_counter_text = get_image_counter_selenium(url, driver)

    if image_counter_text == 'Error al obtener el contador':
        print(f"Fila {row}: Error al obtener el contador (3 intentos).")
    else:
        print(f"Fila {row}: Imperfecciones obtenidas.")
        sheet[f'e{row}'] = image_counter_text #AQUI IBA LA COLUMNA "d" lo cambiaste para revisar el dato de km

    # Guarda el progreso después de cada fila
    workbook.save('/Users/angel1/Desktop/Ops.xlsx')
    print(f"Progreso guardado en la fila {row}.")

# Cierra el navegador y guarda los cambios finales en el libro de Excel
driver.quit()
workbook.save('/Users/angel1/Desktop/Ops.xlsx')
print("Proceso completado. El archivo de Excel ha sido actualizado.")

